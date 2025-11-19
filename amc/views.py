# amc/views.py
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import logging
import json
import csv
import io

from .models import AMCRoutineService, AMCExpiringThisMonth, AMCExpiringLastMonth, AMCExpiringNextMonth, AMC, AMCType
from customer.models import Customer
from items.models import Item
from django.utils import timezone
from datetime import timedelta, datetime, date
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.core.exceptions import ValidationError
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth import get_user_model
from .serializers import AMCCreateSerializer, AMCListSerializer, AMCRoutineServiceSerializer

logger = logging.getLogger(__name__)


def get_customer_json(request, id):
    """API endpoint to get customer details as JSON"""
    try:
        customer = get_object_or_404(Customer, id=id)
        return JsonResponse({
            'id': customer.id,
            'reference_id': customer.reference_id,
            'site_name': customer.site_name,
            'site_address': customer.site_address,
            'job_no': customer.job_no,
            'latitude': str(customer.latitude) if customer.latitude else None,
            'longitude': str(customer.longitude) if customer.longitude else None,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)


def amc_expiring_this_month(request):
    """View for AMCs expiring this month"""
    today = timezone.now().date()
    first_of_month = today.replace(day=1)
    last_day_of_month = (first_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    amcs = AMCExpiringThisMonth.objects.filter(
        end_date__gte=first_of_month,
        end_date__lte=last_day_of_month
    ).select_related('customer').order_by('end_date')
    return render(request, 'amc/amc_expiring_this_month.html', {
        'amcs': amcs,
        'title': 'AMCs Expiring This Month'
    })


def amc_expiring_last_month(request):
    """View for AMCs that expired last month"""
    today = timezone.now().date()
    first_of_month = today.replace(day=1)
    last_month_last_day = first_of_month - timedelta(days=1)
    last_month_first_day = last_month_last_day.replace(day=1)
    amcs = AMCExpiringLastMonth.objects.filter(
        end_date__gte=last_month_first_day,
        end_date__lte=last_month_last_day
    ).select_related('customer').order_by('end_date')
    return render(request, 'amc/amc_expiring_last_month.html', {
        'amcs': amcs,
        'title': 'AMCs Expired Last Month'
    })


def amc_expiring_next_month(request):
    """View for AMCs expiring next month"""
    today = timezone.now().date()
    first_of_month = today.replace(day=1)
    next_month_first = (first_of_month + timedelta(days=32)).replace(day=1)
    next_month_last = (next_month_first + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    amcs = AMCExpiringNextMonth.objects.filter(
        end_date__gte=next_month_first,
        end_date__lte=next_month_last
    ).select_related('customer').order_by('end_date')
    return render(request, 'amc/amc_expiring_next_month.html', {
        'amcs': amcs,
        'title': 'AMCs Expiring Next Month'
    })


# Form views (stubs - may need implementation)
def amc_form(request, pk):
    """Form view for AMC"""
    return JsonResponse({'message': 'AMC form view - to be implemented'}, status=501)


def customer_form(request, pk):
    """Form view for customer"""
    return JsonResponse({'message': 'Customer form view - to be implemented'}, status=501)


# API endpoints for AMC types
@require_http_methods(["GET", "POST"])
@csrf_exempt
def amc_types_list(request):
    """API for listing and creating AMC types"""
    if request.method == 'GET':
        amc_types = AMCType.objects.all().values('id', 'name')
        return JsonResponse(list(amc_types), safe=False)
    elif request.method == 'POST':
        try:
            data = json.loads(request.body) if hasattr(request, 'body') else request.POST
            amc_type = AMCType.objects.create(name=data['name'])
            return JsonResponse({'id': amc_type.id, 'name': amc_type.name}, status=201)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


@require_http_methods(["PUT", "DELETE"])
@csrf_exempt
def amc_types_detail(request, amc_type_id):
    """API for updating/deleting AMC types"""
    try:
        amc_type = get_object_or_404(AMCType, id=amc_type_id)
        if request.method == 'PUT':
            data = json.loads(request.body) if hasattr(request, 'body') else request.POST
            amc_type.name = data.get('name', amc_type.name)
            amc_type.save()
            return JsonResponse({'id': amc_type.id, 'name': amc_type.name})
        elif request.method == 'DELETE':
            amc_type.delete()
            return JsonResponse({'message': 'Deleted successfully'}, status=204)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# API endpoints for payment terms (stubs - payment terms may be disabled)
@require_http_methods(["GET", "POST"])
@csrf_exempt
def payment_terms_list(request):
    """API for listing payment terms"""
    return JsonResponse([], safe=False)


@require_http_methods(["PUT", "DELETE"])
@csrf_exempt
def payment_terms_detail(request, payment_term_id):
    """API for payment terms detail"""
    return JsonResponse({'error': 'Payment terms feature is disabled'}, status=400)


# API to get next AMC reference
@require_http_methods(["GET"])
def get_next_amc_reference(request):
    """Get the next AMC reference ID - matches the logic in AMC.save()"""
    try:
        # Use the same logic as AMC.save() method
        last_amc = AMC.objects.order_by("id").last()
        last_id = 0
        if last_amc and last_amc.reference_id and last_amc.reference_id.startswith("AMC"):
            try:
                # Extract number from reference_id (e.g., "AMC03" -> 3)
                last_id = int(last_amc.reference_id.replace("AMC", ""))
            except ValueError:
                last_id = 0
        next_reference = f"AMC{str(last_id + 1).zfill(2)}"
        # Return both for backward compatibility
        return JsonResponse({
            'next_reference': next_reference,
            'reference_id': next_reference
        })
    except Exception as e:
        logger.error(f"Error generating next AMC reference: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# Mobile app API endpoints
@api_view(['GET'])
@csrf_exempt
def list_amcs_mobile(request):
    """Mobile API to list AMCs"""
    try:
        amcs = AMC.objects.select_related('customer', 'amc_type').all()
        serializer = AMCListSerializer(amcs, many=True)
        return Response({'amcs': serializer.data})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@csrf_exempt
def create_amc_mobile(request):
    """Mobile API to create AMC"""
    try:
        serializer = AMCCreateSerializer(data=request.data)
        if serializer.is_valid():
            amc = serializer.save()
            return Response({
                'message': 'AMC created successfully',
                'amc': {
                    'id': amc.id,
                    'reference_id': amc.reference_id,
                    'customer': amc.customer.site_name if amc.customer else None,
                    'start_date': amc.start_date.strftime('%Y-%m-%d') if amc.start_date else None,
                    'end_date': amc.end_date.strftime('%Y-%m-%d') if amc.end_date else None,
                    'status': amc.status
                }
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@csrf_exempt
def list_amc_types_mobile(request):
    """Mobile API to list AMC types"""
    try:
        amc_types = AMCType.objects.all().values('id', 'name')
        return Response({'amc_types': list(amc_types)})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@csrf_exempt
def create_amc_type_mobile(request):
    """Mobile API to create AMC type"""
    try:
        name = request.data.get('name')
        if not name:
            return Response({'error': 'Name is required'}, status=status.HTTP_400_BAD_REQUEST)
        amc_type = AMCType.objects.create(name=name)
        return Response({
            'message': 'AMC type created successfully',
            'amc_type': {'id': amc_type.id, 'name': amc_type.name}
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@csrf_exempt
def get_employee_routine_services(request):
    """
    Mobile API to get routine services assigned to a specific employee/user.
    
    Query Parameters:
    - user_id (optional): User ID to get services for. If not provided, uses authenticated user.
    - status (optional): Filter by status ('due', 'overdue', 'completed', 'cancelled')
    - start_date (optional): Filter services from this date (YYYY-MM-DD)
    - end_date (optional): Filter services until this date (YYYY-MM-DD)
    - page (optional): Page number for pagination
    - page_size (optional): Number of items per page
    
    Returns:
    - List of routine services assigned to the employee with full details
    """
    try:
        User = get_user_model()
        
        # Get user_id from query params or use authenticated user
        user_id = request.query_params.get('user_id')
        if user_id:
            # Only allow superusers to query other users' services
            if not request.user.is_superuser:
                return Response({
                    'error': 'You do not have permission to view other users\' services'
                }, status=status.HTTP_403_FORBIDDEN)
            try:
                target_user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({
                    'error': 'User not found'
                }, status=status.HTTP_404_NOT_FOUND)
        else:
            # Use authenticated user
            target_user = request.user
        
        # Get all routine services assigned to this user
        queryset = AMCRoutineService.objects.filter(
            employee_assign=target_user
        ).select_related(
            'amc__customer',
            'employee_assign'
        ).order_by('-service_date')
        
        # Apply filters
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        start_date = request.query_params.get('start_date')
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(service_date__gte=start_date_obj)
            except ValueError:
                return Response({
                    'error': 'Invalid start_date format. Use YYYY-MM-DD'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        end_date = request.query_params.get('end_date')
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(service_date__lte=end_date_obj)
            except ValueError:
                return Response({
                    'error': 'Invalid end_date format. Use YYYY-MM-DD'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Pagination
        page = request.query_params.get('page', 1)
        page_size = request.query_params.get('page_size', 20)
        
        try:
            page = int(page)
            page_size = int(page_size)
        except ValueError:
            return Response({
                'error': 'Invalid page or page_size. Must be integers.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Calculate pagination
        total_count = queryset.count()
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        
        services = queryset[start_index:end_index]
        
        # Serialize the services
        serializer = AMCRoutineServiceSerializer(services, many=True)
        
        # Calculate pagination info
        total_pages = (total_count + page_size - 1) // page_size if page_size > 0 else 1
        has_next = page < total_pages
        has_previous = page > 1
        
        response_data = {
            'count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'has_next': has_next,
            'has_previous': has_previous,
            'results': serializer.data,
            'employee': {
                'id': target_user.id,
                'username': target_user.username,
                'email': target_user.email,
                'full_name': f"{target_user.first_name} {target_user.last_name}".strip() or target_user.username
            }
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error retrieving employee routine services: {str(e)}")
        return Response({
            'error': 'Failed to retrieve routine services',
            'detail': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def export_amc_routine_services_xlsx(request, pk):
    """Export AMC routine services to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    
    amc = get_object_or_404(AMC.objects.select_related('customer'), pk=pk)
    routine_services = AMCRoutineService.objects.filter(amc=amc).select_related('employee_assign').order_by('service_date')
    
    # Create workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Routine Services"
    
    # Header row styling
    header_fill = PatternFill(start_color="2D3A6B", end_color="2D3A6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Add AMC info at the top
    worksheet.merge_cells('A1:H1')
    info_cell = worksheet.cell(row=1, column=1, value=f"AMC Routine Services - {amc.reference_id}")
    info_cell.font = Font(bold=True, size=14)
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    worksheet.merge_cells('A2:H2')
    customer_info = f"Customer: {amc.customer.site_name if amc.customer else 'N/A'} | Contract Period: {amc.start_date.strftime('%d/%m/%Y') if amc.start_date else 'N/A'} - {amc.end_date.strftime('%d/%m/%Y') if amc.end_date else 'N/A'}"
    info_cell2 = worksheet.cell(row=2, column=1, value=customer_info)
    info_cell2.font = Font(size=11)
    info_cell2.alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers (row 3)
    headers = ['No', 'Lift Code', 'Date', 'Month', 'Block / Wing', 'Note', 'Assign To', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows (starting from row 4)
    for row_num, service in enumerate(routine_services, 4):
        worksheet.cell(row=row_num, column=1, value=row_num - 3)  # No
        
        # Lift Code (from customer job_no)
        lift_code = amc.customer.job_no if amc.customer and amc.customer.job_no else '—'
        worksheet.cell(row=row_num, column=2, value=lift_code)
        
        # Date
        date_cell = worksheet.cell(row=row_num, column=3)
        if service.service_date:
            date_cell.value = service.service_date
            date_cell.number_format = 'DD/MM/YYYY'
        else:
            date_cell.value = '—'
        
        # Month
        month_name = service.service_date.strftime('%B') if service.service_date else '—'
        worksheet.cell(row=row_num, column=4, value=month_name)
        
        # Block / Wing
        worksheet.cell(row=row_num, column=5, value=service.block_wing or '—')
        
        # Note
        worksheet.cell(row=row_num, column=6, value=service.note or '—')
        
        # Assign To
        if service.employee_assign:
            employee_name = service.employee_assign.get_full_name() or service.employee_assign.username
            worksheet.cell(row=row_num, column=7, value=employee_name)
        else:
            worksheet.cell(row=row_num, column=7, value='Unassigned')
        
        # Status
        status_display = service.get_status_display() if hasattr(service, 'get_status_display') else service.status
        worksheet.cell(row=row_num, column=8, value=status_display)
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # No
        'B': 15,  # Lift Code
        'C': 12,  # Date
        'D': 12,  # Month
        'E': 15,  # Block / Wing
        'F': 30,  # Note
        'G': 20,  # Assign To
        'H': 12,  # Status
    }
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"AMC_Routine_Services_{amc.reference_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    
    return response


def print_routine_service_certificate(request, pk):
    """Generate and download a PDF certificate for a routine service visit"""
    try:
        service = get_object_or_404(
            AMCRoutineService.objects.select_related('amc__customer', 'employee_assign'),
            pk=pk
        )
        
        amc = service.amc
        customer = amc.customer
        
        # Prepare context data (matching complaint PDF format)
        context = {
            'company_name': 'Atom Lifts India Pvt Ltd',
            'address': 'No.87B, Pillayar Koil Street, Mannurpet, Ambattur Indus Estate, Chennai 50., CHENNAI',
            'phone': '9600087456',
            'email': 'admin@atomlifts.com',
            'amc_no': amc.reference_id if amc else 'N/A',
            'service_date': service.service_date.strftime('%d/%m/%Y'),
            'service_month': service.service_date.strftime('%B'),
            'site_name': customer.site_name if customer else 'N/A',
            'site_address': customer.site_address if customer and customer.site_address else 'N/A',
            'assign_to': (
                f"{service.employee_assign.first_name} {service.employee_assign.last_name}".strip()
                or service.employee_assign.username
                if service.employee_assign else "Unassigned"
            ),
            'technician_remark': '',
            'service_provided': '',
            'customer_remark': '',
            'service': '',
            'attend_date_time': '',
            'service_status': service.get_status_display() if service.status != 'due' else 'Due',
        }
        
        # Build PDF (matching complaint PDF format)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
        styles = getSampleStyleSheet()
        story = []
        
        # Header (matching complaint format)
        header_style = ParagraphStyle(
            name='HeaderStyle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1  # Center
        )
        story.append(Paragraph(context['company_name'], header_style))
        story.append(Paragraph(context['address'], styles['Normal']))
        story.append(Paragraph(f"Phone: {context['phone']} | Email: {context['email']}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Certificate Title
        story.append(Paragraph("CERTIFICATE OF ROUTINE SERVICE VISIT", header_style))
        story.append(Spacer(1, 12))
        
        # Service Information (matching complaint format)
        data = [
            ['AMC No.:', context['amc_no']],
            ['Service Date:', context['service_date']],
            ['Service Month:', context['service_month']],
        ]
        table = Table(data, colWidths=[100, 400])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))
        
        # Customer Information (matching complaint format)
        story.append(Paragraph("Customer Information", styles['Heading2']))
        cust_data = [
            ['Site Name:', context['site_name']],
            ['Site Address:', context['site_address']],
            ['Note:', ''],
        ]
        cust_table = Table(cust_data, colWidths=[100, 400])
        cust_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(cust_table)
        story.append(Spacer(1, 12))
        
        # Service Details (matching complaint format)
        story.append(Paragraph("Service Details", styles['Heading2']))
        story.append(Paragraph(f"Assign To: {context['assign_to']}", styles['Normal']))
        story.append(Paragraph(f"Technician Remark: {context['technician_remark']}", styles['Normal']))
        story.append(Paragraph(f"Service Provided: {context['service_provided']}", styles['Normal']))
        story.append(Paragraph(f"Customer Remark: {context['customer_remark']}", styles['Normal']))
        story.append(Paragraph(f"Service: {context['service']}", styles['Normal']))
        story.append(Paragraph(f"Attend Date & Time: {context['attend_date_time']}", styles['Normal']))
        story.append(Paragraph(f"Service Status: {context['service_status']}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Signatures (matching complaint format)
        story.append(Paragraph("Signatures", styles['Heading2']))
        
        # Customer Signature
        story.append(Paragraph("Customer Signature:", styles['Normal']))
        story.append(Paragraph("___________________________", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Technician Signature
        story.append(Paragraph("Technician Signature:", styles['Normal']))
        story.append(Paragraph("___________________________", styles['Normal']))
        story.append(Spacer(1, 12))
        
        story.append(Spacer(1, 12))
        
        # Build and return
        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f'Routine_Service_Certificate_{context["amc_no"]}_{service.service_date.strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.error(f"Error generating routine service certificate PDF: {str(e)}")
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


def bulk_import_view(request):
    """View for bulk importing AMCs from CSV/Excel"""
    if request.method == 'POST':
        try:
            file = request.FILES.get('file')
            if not file:
                messages.error(request, 'Please select a file to upload.')
                return render(request, 'amc/bulk_import.html')
            
            # Check file extension
            file_name = file.name.lower()
            if not (file_name.endswith('.csv') or file_name.endswith('.xlsx') or file_name.endswith('.xls')):
                messages.error(request, 'Please upload a CSV or Excel file (.csv, .xlsx, .xls)')
                return render(request, 'amc/bulk_import.html')
            
            # Read file content
            file_content = file.read()
            
            # Parse CSV
            if file_name.endswith('.csv'):
                try:
                    # Try to decode as UTF-8
                    try:
                        decoded_file = file_content.decode('utf-8')
                    except UnicodeDecodeError:
                        # Try with different encoding
                        decoded_file = file_content.decode('latin-1')
                    
                    csv_reader = csv.DictReader(io.StringIO(decoded_file))
                    # Normalize headers to lowercase, strip whitespace, and replace spaces with underscores
                    rows = []
                    for row in csv_reader:
                        normalized_row = {}
                        for key, value in row.items():
                            # Normalize key: lowercase, strip, and replace spaces/hyphens with underscores
                            if key:
                                normalized_key = key.strip().lower().replace(' ', '_').replace('-', '_')
                            else:
                                normalized_key = ''
                            # Handle None values and convert to string
                            if value is None:
                                normalized_row[normalized_key] = ''
                            else:
                                normalized_row[normalized_key] = str(value) if value else ''
                        rows.append(normalized_row)
                except Exception as e:
                    messages.error(request, f'Error reading CSV file: {str(e)}')
                    return render(request, 'amc/bulk_import.html')
            else:
                # Parse Excel file
                try:
                    import openpyxl
                    workbook = openpyxl.load_workbook(io.BytesIO(file_content))
                    sheet = workbook.active
                    
                    # Get headers from first row, convert to lowercase, strip, and replace spaces with underscores
                    headers = []
                    for cell in sheet[1]:
                        header_value = cell.value
                        if header_value:
                            normalized_header = str(header_value).strip().lower().replace(' ', '_').replace('-', '_')
                            headers.append(normalized_header)
                        else:
                            headers.append('')
                    
                    rows = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Check if row has any non-empty values
                        if any(cell is not None and str(cell).strip() for cell in row if cell is not None):
                            # Create dict with lowercase keys and handle None values
                            row_dict = {}
                            for i, cell_value in enumerate(row):
                                if i < len(headers) and headers[i]:
                                    # Convert None to empty string, then to string
                                    if cell_value is None:
                                        row_dict[headers[i]] = ''
                                    else:
                                        # Handle datetime/date objects from Excel (convert to YYYY-MM-DD format)
                                        if isinstance(cell_value, (datetime, date)):
                                            row_dict[headers[i]] = cell_value.strftime('%Y-%m-%d')
                                        else:
                                            row_dict[headers[i]] = str(cell_value)
                            if row_dict:  # Only add if dict is not empty
                                rows.append(row_dict)
                except ImportError:
                    messages.error(request, 'openpyxl library is required for Excel files. Please install it: pip install openpyxl')
                    return render(request, 'amc/bulk_import.html')
                except Exception as e:
                    messages.error(request, f'Error reading Excel file: {str(e)}')
                    return render(request, 'amc/bulk_import.html')
            
            if not rows:
                messages.error(request, 'The file appears to be empty or has no data rows.')
                return render(request, 'amc/bulk_import.html')
            
            # Process rows and create AMCs
            success_count = 0
            error_count = 0
            errors = []
            
            # Helper function to parse dates in various formats
            def parse_date(date_str):
                """Parse date string in YYYY-MM-DD format or common Excel formats"""
                if not date_str or not date_str.strip():
                    return None
                date_str = date_str.strip()
                
                # Try YYYY-MM-DD format first (expected format)
                try:
                    return datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    pass
                
                # Try other common formats
                date_formats = [
                    '%Y/%m/%d',
                    '%d/%m/%Y',
                    '%m/%d/%Y',
                    '%d-%m-%Y',
                    '%m-%d-%Y',
                ]
                for fmt in date_formats:
                    try:
                        return datetime.strptime(date_str, fmt).date()
                    except ValueError:
                        continue
                
                # If all parsing fails, return None (will be caught by error handling)
                return None
            
            # Helper functions for parsing numeric values
            def parse_int(value, default=None):
                """Parse integer value, return None if empty/invalid"""
                if value is None or value == '':
                    return default
                try:
                    return int(value) if str(value).strip() else default
                except (ValueError, TypeError):
                    return default

            def parse_decimal(value, default=None):
                """Parse decimal value, return None if empty/invalid"""
                if value is None or value == '':
                    return default
                try:
                    return float(value) if str(value).strip() else default
                except (ValueError, TypeError):
                    return default
            
            for idx, row in enumerate(rows, start=2):  # Start from 2 (1 is header)
                try:
                    # Map CSV columns to model fields - handle None values and empty strings
                    # Headers are normalized to lowercase with underscores
                    
                    # Required fields (from add_amc_custom requirements)
                    customer_value = row.get('customer', '') or row.get('customer_value', '') or ''
                    customer_value = str(customer_value).strip() if customer_value else ''
                    
                    # Handle both 'start_date' and 'start_date_str' column names
                    start_date_raw = row.get('start_date', '') or row.get('start_date_str', '') or ''
                    start_date = str(start_date_raw).strip() if start_date_raw else ''
                    
                    # Handle both 'end_date' and 'end_date_str' column names
                    end_date_raw = row.get('end_date', '') or row.get('end_date_str', '') or ''
                    end_date = str(end_date_raw).strip() if end_date_raw else ''
                    
                    # Validate required fields (same as add_amc_custom)
                    if not customer_value:
                        errors.append(f'Row {idx}: Customer is required.')
                        error_count += 1
                        continue
                    
                    if not start_date:
                        errors.append(f'Row {idx}: Start date is required.')
                        error_count += 1
                        continue
                    
                    if not end_date:
                        errors.append(f'Row {idx}: End date is required.')
                        error_count += 1
                        continue
                    
                    # Get customer by site_name
                    customer = Customer.objects.filter(site_name=customer_value).first()
                    if not customer:
                        errors.append(f'Row {idx}: Customer "{customer_value}" not found. Please use an existing customer site name.')
                        error_count += 1
                        continue
                    
                    # Parse dates
                    start_date_parsed = parse_date(start_date)
                    if start_date_parsed is None:
                        errors.append(f'Row {idx}: Invalid start date format. Please use YYYY-MM-DD format.')
                        error_count += 1
                        continue
                    
                    end_date_parsed = parse_date(end_date)
                    if end_date_parsed is None:
                        errors.append(f'Row {idx}: Invalid end date format. Please use YYYY-MM-DD format.')
                        error_count += 1
                        continue
                    
                    # Validate date order
                    if start_date_parsed >= end_date_parsed:
                        errors.append(f'Row {idx}: Start date must be before end date.')
                        error_count += 1
                        continue
                    
                    # Optional fields
                    amc_type_value = row.get('amc_type', '') or row.get('amc_type_value', '') or ''
                    amc_type_value = str(amc_type_value).strip() if amc_type_value else ''
                    
                    amc_type = None
                    if amc_type_value:
                        amc_type = AMCType.objects.filter(name=amc_type_value).first()
                        if not amc_type:
                            # Create if doesn't exist
                            amc_type = AMCType.objects.create(name=amc_type_value)
                    
                    payment_terms_value = row.get('payment_terms', '') or row.get('payment_terms_value', '') or ''
                    payment_terms_value = str(payment_terms_value).strip() if payment_terms_value else ''
                    
                    payment_terms = None
                    if payment_terms_value:
                        from .models import PaymentTerms
                        payment_terms, created = PaymentTerms.objects.get_or_create(name=payment_terms_value)
                    
                    amc_service_item_value = row.get('amc_service_item', '') or row.get('amc_service_item_value', '') or ''
                    amc_service_item_value = str(amc_service_item_value).strip() if amc_service_item_value else ''
                    
                    amc_service_item = None
                    if amc_service_item_value:
                        amc_service_item = Item.objects.filter(name=amc_service_item_value).first()
                        if not amc_service_item:
                            errors.append(f'Row {idx}: AMC Service Item "{amc_service_item_value}" not found.')
                            error_count += 1
                            continue
                    
                    # Validate contract generation (same as add_amc_custom)
                    generate_contract = row.get('is_generate_contract', '')
                    if isinstance(generate_contract, str):
                        generate_contract = generate_contract.lower() in ('true', 'on', '1', 'yes')
                    else:
                        generate_contract = bool(generate_contract)
                    
                    if generate_contract and not amc_service_item:
                        errors.append(f'Row {idx}: Please select an AMC Service Item when generating contract.')
                        error_count += 1
                        continue
                    
                    # Parse numeric fields
                    invoice_frequency = row.get('invoice_frequency', 'annually') or 'annually'
                    no_of_services = parse_int(row.get('no_of_services'))
                    price = parse_decimal(row.get('price'), default=0)
                    no_of_lifts = parse_int(row.get('no_of_lifts'), default=0)
                    gst_percentage = parse_decimal(row.get('gst_percentage'), default=0)
                    total_amount_paid = parse_decimal(row.get('total_amount_paid'), default=0)
                    geo_latitude = parse_decimal(row.get('geo_latitude'))
                    geo_longitude = parse_decimal(row.get('geo_longitude'))
                    
                    # Text fields
                    equipment_no = row.get('equipment_no', '') or ''
                    equipment_no = str(equipment_no).strip() if equipment_no else ''
                    
                    latitude = row.get('latitude', '') or ''
                    latitude = str(latitude).strip() if latitude else ''
                    
                    notes = row.get('notes', '') or ''
                    notes = str(notes).strip() if notes else ''
                    
                    # Create AMC (same structure as add_amc_custom)
                    amc = AMC.objects.create(
                        customer=customer,
                        invoice_frequency=invoice_frequency,
                        amc_type=amc_type,
                        start_date=start_date_parsed,
                        end_date=end_date_parsed,
                        equipment_no=equipment_no,
                        latitude=latitude,
                        geo_latitude=geo_latitude,
                        geo_longitude=geo_longitude,
                        notes=notes,
                        is_generate_contract=generate_contract,
                        no_of_services=no_of_services,
                        amc_service_item=amc_service_item,
                        price=price,
                        no_of_lifts=no_of_lifts,
                        gst_percentage=gst_percentage,
                        total_amount_paid=total_amount_paid,
                        payment_terms=payment_terms,
                    )
                    
                    # Validate and save (uses full_clean which applies all model validations)
                    try:
                        amc.full_clean()
                        amc.save()
                        success_count += 1
                    except ValidationError as e:
                        # Handle validation errors
                        if e.message_dict:
                            error_fields = ['customer', 'start_date', 'end_date']
                            error_msg = None
                            for field in error_fields:
                                if field in e.message_dict:
                                    error_msg = f"Row {idx}: {e.message_dict[field][0]}"
                                    break
                            if not error_msg:
                                error_msg = f"Row {idx}: {list(e.message_dict.values())[0][0]}"
                        else:
                            error_msg = f"Row {idx}: {str(e)}"
                        errors.append(error_msg)
                        error_count += 1
                        continue
                    except Exception as e:
                        # Handle unique constraint violations and other database errors
                        error_str = str(e).lower()
                        if 'unique' in error_str or 'duplicate' in error_str or 'already exists' in error_str:
                            errors.append(f'Row {idx}: Duplicate entry - {str(e)}')
                        else:
                            errors.append(f'Row {idx}: {str(e)}')
                        error_count += 1
                        continue
                        
                except Exception as e:
                    errors.append(f'Row {idx}: Unexpected error - {str(e)}')
                    error_count += 1
                    continue
            
            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} AMC(s).')
            if error_count > 0:
                error_message = f'Failed to import {error_count} row(s).'
                if errors:
                    error_message += ' Errors: ' + '; '.join(errors[:10])  # Show first 10 errors
                    if len(errors) > 10:
                        error_message += f' ... and {len(errors) - 10} more error(s).'
                messages.error(request, error_message)
            
            return render(request, 'amc/bulk_import.html')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return render(request, 'amc/bulk_import.html')
    
    # GET request - render form
    return render(request, 'amc/bulk_import.html')
