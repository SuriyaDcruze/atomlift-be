from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count
import json
from datetime import datetime, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from complaints.models import Complaint
from invoice.models import Invoice
from PaymentReceived.models import PaymentReceived
from Quotation.models import Quotation
from amc.models import AMC
from customer.models import Customer


@login_required
def complaints_report(request):
    """Complaints Report View"""
    view_mode = request.GET.get('view')
    # Get filter parameters
    period = request.GET.get('period', 'ALL TIME')
    customer_filter = request.GET.get('customer', 'ALL')
    by_filter = request.GET.get('by', 'ALL')
    status_filter = request.GET.get('status', 'ALL')
    
    # Base queryset
    complaints = Complaint.objects.all().select_related('customer', 'assign_to', 'complaint_type', 'priority')
    
    # Apply filters
    if period == 'LAST_WEEK':
        week_ago = datetime.now() - timedelta(days=7)
        complaints = complaints.filter(date__gte=week_ago)
    elif period == 'LAST_MONTH':
        month_ago = datetime.now() - timedelta(days=30)
        complaints = complaints.filter(date__gte=month_ago)
    
    if customer_filter != 'ALL':
        complaints = complaints.filter(customer__site_name=customer_filter)
    
    if status_filter != 'ALL':
        # Map display names to model values
        status_map = {
            'Open': 'open',
            'In Progress': 'in_progress',
            'Closed': 'closed',
        }
        status_value = status_map.get(status_filter, status_filter.lower().replace(' ', '_'))
        complaints = complaints.filter(status=status_value)
    
    # Get customer list for filter dropdown
    customers = Customer.objects.all().values_list('site_name', flat=True).distinct()
    
    if view_mode == 'graph':
        by_status = complaints.values('status').annotate(count=Count('id')).order_by()
        # Map status values to display names
        status_display_map = {
            'open': 'Open',
            'in_progress': 'In Progress',
            'closed': 'Closed',
        }
        labels = [status_display_map.get(row['status'], row['status'] or 'Unknown') for row in by_status]
        data = [row['count'] for row in by_status]
        context = {
            'graph_title': 'Complaints by Status',
            'labels_json': json.dumps(labels),
            'datasets_json': json.dumps([
                {
                    'label': 'Count',
                    'data': data,
                    'backgroundColor': ['#60a5fa', '#34d399', '#fbbf24', '#a78bfa', '#f87171'][:len(data)],
                }
            ]),
            'chart_type': 'bar',
        }
        return render(request, 'reports/graph_report.html', context)

    context = {
        'complaints': complaints,
        'customers': customers,
        'selected_period': period,
        'selected_customer': customer_filter,
        'selected_by': by_filter,
        'selected_status': status_filter,
    }
    return render(request, 'reports/complaints_report.html', context)


@login_required
def invoice_report(request):
    """Invoice Report View"""
    view_mode = request.GET.get('view')
    # Get filter parameters
    period = request.GET.get('period', 'ALL TIME')
    customer_filter = request.GET.get('customer', 'ALL')
    by_filter = request.GET.get('by', 'ALL')
    status_filter = request.GET.get('status', 'ALL')
    
    # Base queryset
    invoices = Invoice.objects.all().select_related('customer', 'amc_type')
    
    # Apply filters
    if period == 'CURRENT MONTH':
        invoices = invoices.filter(
            start_date__month=datetime.now().month,
            start_date__year=datetime.now().year
        )
    
    if customer_filter != 'ALL':
        invoices = invoices.filter(customer__site_name=customer_filter)
    
    if status_filter != 'ALL':
        invoices = invoices.filter(status=status_filter)
    
    # Get customer list for filter dropdown
    customers = Customer.objects.all().values_list('site_name', flat=True).distinct()
    
    if view_mode == 'graph':
        by_status = invoices.values('status').annotate(count=Count('id')).order_by()
        labels = [row['status'] or 'Unknown' for row in by_status]
        data = [row['count'] for row in by_status]
        context = {
            'graph_title': 'Invoices by Status',
            'labels_json': json.dumps(labels),
            'datasets_json': json.dumps([
                {
                    'label': 'Count',
                    'data': data,
                    'backgroundColor': ['#60a5fa', '#34d399', '#fbbf24', '#a78bfa', '#f87171'][:len(data)],
                }
            ]),
            'chart_type': 'doughnut',
        }
        return render(request, 'reports/graph_report.html', context)

    context = {
        'invoices': invoices,
        'customers': customers,
        'selected_period': period,
        'selected_customer': customer_filter,
        'selected_by': by_filter,
        'selected_status': status_filter,
    }
    return render(request, 'reports/invoice_report.html', context)


@login_required
def payment_report(request):
    """Payment Report View"""
    view_mode = request.GET.get('view')
    # Get filter parameters
    customer_filter = request.GET.get('customer', '')
    status_filter = request.GET.get('status', 'ALL')
    payment_mode = request.GET.get('payment_mode', 'MONTH')
    period = request.GET.get('period', 'ALL')
    month = request.GET.get('month', datetime.now().strftime('%B'))
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Base queryset
    payments = PaymentReceived.objects.all().select_related('customer', 'invoice')
    
    # Apply filters
    if customer_filter:
        payments = payments.filter(customer__site_name=customer_filter)
    
    if start_date and end_date:
        payments = payments.filter(
            date__gte=start_date,
            date__lte=end_date
        )
    
    # Get customer list for filter dropdown
    customers = Customer.objects.all().values_list('site_name', flat=True).distinct()
    
    if view_mode == 'graph':
        # Aggregate by month and sum amount
        qs = payments
        # default to last 12 months if no range
        if not (start_date and end_date):
            twelve_months_ago = datetime.now() - timedelta(days=365)
            qs = qs.filter(date__gte=twelve_months_ago)
        totals = qs.values('date__year', 'date__month').annotate(total=Sum('amount')).order_by('date__year', 'date__month')
        labels = [f"{row['date__year']}-{str(row['date__month']).zfill(2)}" for row in totals]
        data = [float(row['total'] or 0) for row in totals]
        
        # Ensure we have at least empty arrays if no data
        if not labels:
            labels = []
        if not data:
            data = []
        
        context = {
            'graph_title': 'Payments by Month (Total Amount)',
            'labels_json': json.dumps(labels),
            'datasets_json': json.dumps([
                {
                    'label': 'Amount',
                    'data': data,
                    'borderColor': '#10b981',
                    'backgroundColor': 'rgba(16,185,129,0.2)',
                    'fill': True,
                    'tension': 0.4,
                }
            ]),
            'chart_type': 'line',
        }
        return render(request, 'reports/graph_report.html', context)

    context = {
        'payments': payments,
        'customers': customers,
        'selected_customer': customer_filter,
        'selected_status': status_filter,
        'selected_payment_mode': payment_mode,
        'selected_period': period,
        'selected_month': month,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/payment_report.html', context)


@login_required
def quotation_report(request):
    """Quotation Report View"""
    view_mode = request.GET.get('view')
    # Get filter parameters
    period = request.GET.get('period', 'ALL TIME')
    customer_filter = request.GET.get('customer', 'ALL')
    by_filter = request.GET.get('by', 'ALL')
    status_filter = request.GET.get('status', 'ALL')
    
    # Base queryset
    quotations = Quotation.objects.all().select_related('customer', 'amc_type', 'sales_service_executive')
    
    # Apply filters
    if period == 'Today':
        quotations = quotations.filter(date=datetime.now().date())
    elif period == 'This Week':
        week_start = datetime.now() - timedelta(days=datetime.now().weekday())
        quotations = quotations.filter(date__gte=week_start)
    elif period == 'This Month':
        quotations = quotations.filter(
            date__month=datetime.now().month,
            date__year=datetime.now().year
        )
    
    if customer_filter != 'ALL':
        quotations = quotations.filter(customer__site_name=customer_filter)
    
    if status_filter != 'ALL':
        quotations = quotations.filter(type=status_filter)

    # Get customer list for filter dropdown
    customers = Customer.objects.all().values_list('site_name', flat=True).distinct()
    
    if view_mode == 'graph':
        by_status = quotations.values('type').annotate(count=Count('id')).order_by()
        labels = [row['type'] or 'Unknown' for row in by_status]
        data = [row['count'] for row in by_status]
        context = {
            'graph_title': 'Quotations by Type',
            'labels_json': json.dumps(labels),
            'datasets_json': json.dumps([
                {
                    'label': 'Count',
                    'data': data,
                    'backgroundColor': ['#60a5fa', '#34d399', '#fbbf24', '#a78bfa', '#f87171'][:len(data)],
                }
            ]),
            'chart_type': 'pie',
        }
        return render(request, 'reports/graph_report.html', context)

    context = {
        'quotations': quotations,
        'customers': customers,
        'selected_period': period,
        'selected_customer': customer_filter,
        'selected_by': by_filter,
        'selected_status': status_filter,
    }
    return render(request, 'reports/quotation_report.html', context)


@login_required
def routine_service_report(request):
    """Routine Service Report View"""
    view_mode = request.GET.get('view')
    # Get filter parameters
    customer_filter = request.GET.get('customer', '')
    by_filter = request.GET.get('by', 'DATE')
    status_filter = request.GET.get('status', 'ALL')
    amc_type_filter = request.GET.get('amc_types', '')
    period = request.GET.get('period', 'MONTH')
    month = request.GET.get('month', datetime.now().strftime('%B'))
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Base queryset
    amc_records = AMC.objects.all().select_related('customer', 'amctype')
    
    # Apply filters
    if customer_filter:
        amc_records = amc_records.filter(customer__site_name=customer_filter)
    
    if amc_type_filter:
        amc_records = amc_records.filter(amctype__amc_type_name=amc_type_filter)
    
    if status_filter != 'ALL':
        amc_records = amc_records.filter(status=status_filter)
    
    if start_date and end_date:
        amc_records = amc_records.filter(
            contract_start_date__gte=start_date,
            contract_start_date__lte=end_date
        )
    
    # Get customer list for filter dropdown
    customers = Customer.objects.all().values_list('site_name', flat=True).distinct()
    
    if view_mode == 'graph':
        by_type = amc_records.values('amctype__amc_type_name').annotate(count=Count('id')).order_by()
        labels = [row['amctype__amc_type_name'] or 'Unknown' for row in by_type]
        data = [row['count'] for row in by_type]
        context = {
            'graph_title': 'Routine Services by AMC Type',
            'labels_json': json.dumps(labels),
            'datasets_json': json.dumps([
                {
                    'label': 'Count',
                    'data': data,
                    'backgroundColor': ['#60a5fa', '#34d399', '#fbbf24', '#a78bfa', '#f87171'][:len(data)],
                }
            ]),
            'chart_type': 'bar',
        }
        return render(request, 'reports/graph_report.html', context)

    context = {
        'amc_records': amc_records,
        'customers': customers,
        'selected_customer': customer_filter,
        'selected_by': by_filter,
        'selected_status': status_filter,
        'selected_amc_type': amc_type_filter,
        'selected_period': period,
        'selected_month': month,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/routine_service_report.html', context)


@login_required
def amc_report(request):
    """AMC Report View"""
    view_mode = request.GET.get('view')

    # Filters
    period = request.GET.get('period', 'ALL TIME')
    customer_filter = request.GET.get('customer', 'ALL')
    status_filter = request.GET.get('status', 'ALL')
    amc_type_filter = request.GET.get('amc_type', 'ALL')

    # Base queryset
    amcs = AMC.objects.all().select_related('customer', 'amc_type')

    # Apply period filter
    if period == 'CURRENT MONTH':
        amcs = amcs.filter(
            start_date__month=datetime.now().month,
            start_date__year=datetime.now().year
        )
    elif period == 'LAST 90 DAYS':
        amcs = amcs.filter(start_date__gte=datetime.now() - timedelta(days=90))

    # Apply other filters
    if customer_filter != 'ALL' and customer_filter:
        amcs = amcs.filter(customer__site_name=customer_filter)

    if status_filter != 'ALL' and status_filter:
        amcs = amcs.filter(status=status_filter)

    if amc_type_filter != 'ALL' and amc_type_filter:
        amcs = amcs.filter(amc_type__name=amc_type_filter)

    # Dropdown data
    customers = Customer.objects.all().values_list('site_name', flat=True).distinct()
    from amc.models import AMCType
    amc_types = AMCType.objects.all().values_list('name', flat=True)

    if view_mode == 'graph':
        # Graph per customer: stacked bar of Contract Amount, Total Paid, Amount Due
        per_customer_qs = amcs.values('customer__site_name').annotate(
            total_amount=Sum('contract_amount'),
            total_paid=Sum('total_amount_paid'),
            total_due=Sum('amount_due'),
        ).order_by('customer__site_name')

        # Pagination across customers (10 per page)
        page_size = 10
        try:
            page = int(request.GET.get('page', '1'))
        except ValueError:
            page = 1
        if page < 1:
            page = 1

        per_customer = list(per_customer_qs)
        total_count = len(per_customer)
        total_pages = (total_count + page_size - 1) // page_size if total_count else 1
        if page > total_pages:
            page = total_pages
        start_index = (page - 1) * page_size
        end_index = start_index + page_size
        current_rows = per_customer[start_index:end_index]

        labels = [row['customer__site_name'] or 'Unknown' for row in current_rows]
        dataset_amount = [float(row['total_amount'] or 0) for row in current_rows]
        dataset_paid = [float(row['total_paid'] or 0) for row in current_rows]
        dataset_due = [float(row['total_due'] or 0) for row in current_rows]

        datasets = [
            {
                'label': 'Contract Amount',
                'data': dataset_amount,
                'backgroundColor': '#60a5fa'
            },
            {
                'label': 'Total Paid',
                'data': dataset_paid,
                'backgroundColor': '#34d399'
            },
            {
                'label': 'Amount Due',
                'data': dataset_due,
                'backgroundColor': '#f87171'
            }
        ]

        # Build prev/next links preserving filters
        def build_url(target_page: int) -> str:
            query = request.GET.copy()
            query['view'] = 'graph'
            query['page'] = str(target_page)
            return f"{request.path}?{query.urlencode()}"

        has_prev = page > 1
        has_next = page < total_pages
        prev_url = build_url(page - 1) if has_prev else ''
        next_url = build_url(page + 1) if has_next else ''

        context = {
            'graph_title': 'AMC Financials by Customer',
            'labels_json': json.dumps(labels),
            'datasets_json': json.dumps(datasets),
            'chart_type': 'bar',
            'stacked': True,
            # pagination meta for template controls
            'total_count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': total_pages,
            'start_index_display': (start_index + 1) if total_count else 0,
            'end_index_display': min(end_index, total_count),
            'has_prev': has_prev,
            'has_next': has_next,
            'prev_url': prev_url,
            'next_url': next_url,
        }
        return render(request, 'reports/graph_report.html', context)

    context = {
        'amcs': amcs,
        'customers': customers,
        'amc_types': amc_types,
        'selected_period': period,
        'selected_customer': customer_filter,
        'selected_status': status_filter,
        'selected_amc_type': amc_type_filter,
    }
    return render(request, 'reports/amc_report.html', context)


@login_required
def export_complaints_csv(request):
    """Export Complaints to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="complaints_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Complaint No', 'Date', 'Customer', 'Contact Person', 'Mobile',
        'Type', 'Problem', 'Resolution', 'Assigned To', 'Priority'
    ])
    
    complaints = Complaint.objects.all().select_related('customer', 'assign_to', 'complaint_type', 'priority')
    for complaint in complaints:
        writer.writerow([
            complaint.reference or complaint.id,
            # Use ISO date format so Excel/Sheets parse reliably
            complaint.date.strftime('%Y-%m-%d') if complaint.date else '',
            complaint.customer.site_name if complaint.customer else 'Unknown',
            complaint.contact_person_name or 'N/A',
            complaint.contact_person_mobile or 'N/A',
            complaint.complaint_type.name if complaint.complaint_type else 'General',
            complaint.subject or 'N/A',
            complaint.solution or 'Pending',
            complaint.assign_to.get_full_name() if complaint.assign_to else 'Unassigned',
            complaint.priority.name if complaint.priority else 'Open',
        ])
    
    return response


@login_required
def export_invoices_csv(request):
    """Export Invoices to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="invoice_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Invoice ID', 'Customer', 'AMC Type', 'Invoice Date', 'Due Date',
        'Discount', 'Payment Term', 'Status'
    ])
    
    invoices = Invoice.objects.all().select_related('customer', 'amc_type')
    for invoice in invoices:
        writer.writerow([
            invoice.reference_id or invoice.id,
            invoice.customer.site_name if invoice.customer else 'N/A',
            invoice.amc_type.name if invoice.amc_type else 'N/A',
            invoice.start_date.strftime('%Y-%m-%d') if invoice.start_date else '',
            invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '',
            f'{invoice.discount}%' if invoice.discount else '0%',
            invoice.get_payment_term_display() if invoice.payment_term else 'N/A',
            invoice.get_status_display() if invoice.status else 'Open',
        ])
    
    return response


@login_required
def export_quotations_csv(request):
    """Export Quotations to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="quotation_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Reference ID', 'Date', 'Customer', 'AMC Type', 'Quotation Type',
        'Sales Executive', 'Year of Make'
    ])
    
    quotations = Quotation.objects.all().select_related('customer', 'amc_type', 'sales_service_executive')
    for quotation in quotations:
        writer.writerow([
            quotation.reference_id or quotation.id,
            quotation.date.strftime('%Y-%m-%d') if quotation.date else '',
            quotation.customer.site_name if quotation.customer else 'N/A',
            quotation.amc_type.name if quotation.amc_type else 'N/A',
            quotation.type or 'N/A',
            quotation.sales_service_executive.get_full_name() if quotation.sales_service_executive else 'N/A',
            quotation.year_of_make or 'N/A',
        ])
    
    return response


@login_required
def export_complaints_xlsx(request):
    """Export Complaints to XLSX"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Complaints Report"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Complaint No', 'Date', 'Customer', 'Contact Person', 'Mobile',
        'Type', 'Problem', 'Resolution', 'Assigned To', 'Priority'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    complaints = Complaint.objects.all().select_related('customer', 'assign_to', 'complaint_type', 'priority')
    for row_num, complaint in enumerate(complaints, 2):
        worksheet.cell(row=row_num, column=1, value=complaint.reference or str(complaint.id))
        date_cell = worksheet.cell(row=row_num, column=2)
        if complaint.date:
            date_cell.value = complaint.date
            date_cell.number_format = 'DD-MM-YYYY'
        else:
            date_cell.value = ''
        worksheet.cell(row=row_num, column=3, value=complaint.customer.site_name if complaint.customer else 'Unknown')
        worksheet.cell(row=row_num, column=4, value=complaint.contact_person_name or 'N/A')
        worksheet.cell(row=row_num, column=5, value=complaint.contact_person_mobile or 'N/A')
        worksheet.cell(row=row_num, column=6, value=complaint.complaint_type.name if complaint.complaint_type else 'General')
        worksheet.cell(row=row_num, column=7, value=complaint.subject or 'N/A')
        worksheet.cell(row=row_num, column=8, value=complaint.solution or 'Pending')
        worksheet.cell(row=row_num, column=9, value=complaint.assign_to.get_full_name() if complaint.assign_to else 'Unassigned')
        worksheet.cell(row=row_num, column=10, value=complaint.priority.name if complaint.priority else 'Open')
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="complaints_report.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def export_invoices_xlsx(request):
    """Export Invoices to XLSX"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoice Report"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Invoice ID', 'Customer', 'AMC Type', 'Invoice Date', 'Due Date',
        'Discount', 'Payment Term', 'Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    invoices = Invoice.objects.all().select_related('customer', 'amc_type')
    for row_num, invoice in enumerate(invoices, 2):
        worksheet.cell(row=row_num, column=1, value=invoice.reference_id or str(invoice.id))
        worksheet.cell(row=row_num, column=2, value=invoice.customer.site_name if invoice.customer else 'N/A')
        worksheet.cell(row=row_num, column=3, value=invoice.amc_type.name if invoice.amc_type else 'N/A')
        start_cell = worksheet.cell(row=row_num, column=4)
        if invoice.start_date:
            start_cell.value = invoice.start_date
            start_cell.number_format = 'DD-MM-YYYY'
        else:
            start_cell.value = ''
        due_cell = worksheet.cell(row=row_num, column=5)
        if invoice.due_date:
            due_cell.value = invoice.due_date
            due_cell.number_format = 'DD-MM-YYYY'
        else:
            due_cell.value = ''
        worksheet.cell(row=row_num, column=6, value=f'{invoice.discount}%' if invoice.discount else '0%')
        worksheet.cell(row=row_num, column=7, value=invoice.get_payment_term_display() if invoice.payment_term else 'N/A')
        worksheet.cell(row=row_num, column=8, value=invoice.get_status_display() if invoice.status else 'Open')
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="invoice_report.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def export_quotations_xlsx(request):
    """Export Quotations to XLSX"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Quotation Report"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Reference ID', 'Date', 'Customer', 'AMC Type', 'Quotation Type',
        'Sales Executive', 'Year of Make'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    quotations = Quotation.objects.all().select_related('customer', 'amc_type', 'sales_service_executive')
    for row_num, quotation in enumerate(quotations, 2):
        worksheet.cell(row=row_num, column=1, value=quotation.reference_id or str(quotation.id))
        date_cell = worksheet.cell(row=row_num, column=2)
        if quotation.date:
            date_cell.value = quotation.date
            date_cell.number_format = 'DD-MM-YYYY'
        else:
            date_cell.value = ''
        worksheet.cell(row=row_num, column=3, value=quotation.customer.site_name if quotation.customer else 'N/A')
        worksheet.cell(row=row_num, column=4, value=quotation.amc_type.name if quotation.amc_type else 'N/A')
        worksheet.cell(row=row_num, column=5, value=quotation.type or 'N/A')
        worksheet.cell(row=row_num, column=6, value=quotation.sales_service_executive.get_full_name() if quotation.sales_service_executive else 'N/A')
        worksheet.cell(row=row_num, column=7, value=quotation.year_of_make or 'N/A')
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="quotation_report.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def export_payments_csv(request):
    """Export Payments to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="payment_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Payment Number', 'Date', 'Customer', 'Invoice Reference',
        'Amount', 'Payment Type', 'Tax Deducted'
    ])
    
    payments = PaymentReceived.objects.all().select_related('customer', 'invoice')
    for payment in payments:
        writer.writerow([
            payment.payment_number or payment.id,
            payment.date.strftime('%Y-%m-%d') if payment.date else '',
            payment.customer.site_name if payment.customer else 'N/A',
            payment.invoice.reference_id if payment.invoice else 'N/A',
            f'INR {payment.amount or 0.00}',
            payment.get_payment_type_display() if payment.payment_type else 'N/A',
            payment.get_tax_deducted_display() if payment.tax_deducted else 'No',
        ])
    
    return response


@login_required
def export_payments_xlsx(request):
    """Export Payments to XLSX"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Payment Report"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Payment Number', 'Date', 'Customer', 'Invoice Reference',
        'Amount', 'Payment Type', 'Tax Deducted'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    payments = PaymentReceived.objects.all().select_related('customer', 'invoice')
    for row_num, payment in enumerate(payments, 2):
        worksheet.cell(row=row_num, column=1, value=payment.payment_number or str(payment.id))
        date_cell = worksheet.cell(row=row_num, column=2)
        if payment.date:
            date_cell.value = payment.date
            date_cell.number_format = 'DD-MM-YYYY'
        else:
            date_cell.value = ''
        worksheet.cell(row=row_num, column=3, value=payment.customer.site_name if payment.customer else 'N/A')
        worksheet.cell(row=row_num, column=4, value=payment.invoice.reference_id if payment.invoice else 'N/A')
        worksheet.cell(row=row_num, column=5, value=f'INR {payment.amount or 0.00}')
        worksheet.cell(row=row_num, column=6, value=payment.get_payment_type_display() if payment.payment_type else 'N/A')
        worksheet.cell(row=row_num, column=7, value=payment.get_tax_deducted_display() if payment.tax_deducted else 'No')
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payment_report.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def export_amc_csv(request):
    """Export AMC to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="amc_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Reference ID', 'Customer', 'AMC Type', 'Start Date', 'End Date',
        'Contract Amount', 'Total Paid', 'Amount Due', 'Status'
    ])
    
    amcs = AMC.objects.all().select_related('customer', 'amc_type')
    for amc in amcs:
        writer.writerow([
            amc.reference_id or amc.id,
            amc.customer.site_name if amc.customer else 'N/A',
            amc.amc_type.name if amc.amc_type else 'N/A',
            amc.start_date.strftime('%Y-%m-%d') if amc.start_date else '',
            amc.end_date.strftime('%Y-%m-%d') if amc.end_date else '',
            f'INR {amc.contract_amount or 0.00}',
            f'INR {amc.total_amount_paid or 0.00}',
            f'INR {amc.amount_due or 0.00}',
            amc.get_status_display() if amc.status else 'Active',
        ])
    
    return response


@login_required
def export_amc_xlsx(request):
    """Export AMC to XLSX"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "AMC Report"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Reference ID', 'Customer', 'AMC Type', 'Start Date', 'End Date',
        'Contract Amount', 'Total Paid', 'Amount Due', 'Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    amcs = AMC.objects.all().select_related('customer', 'amc_type')
    for row_num, amc in enumerate(amcs, 2):
        worksheet.cell(row=row_num, column=1, value=amc.reference_id or str(amc.id))
        worksheet.cell(row=row_num, column=2, value=amc.customer.site_name if amc.customer else 'N/A')
        worksheet.cell(row=row_num, column=3, value=amc.amc_type.name if amc.amc_type else 'N/A')
        start_cell = worksheet.cell(row=row_num, column=4)
        if amc.start_date:
            start_cell.value = amc.start_date
            start_cell.number_format = 'DD-MM-YYYY'
        else:
            start_cell.value = ''
        end_cell = worksheet.cell(row=row_num, column=5)
        if amc.end_date:
            end_cell.value = amc.end_date
            end_cell.number_format = 'DD-MM-YYYY'
        else:
            end_cell.value = ''
        worksheet.cell(row=row_num, column=6, value=f'INR {amc.contract_amount or 0.00}')
        worksheet.cell(row=row_num, column=7, value=f'INR {amc.total_amount_paid or 0.00}')
        worksheet.cell(row=row_num, column=8, value=f'INR {amc.amount_due or 0.00}')
        worksheet.cell(row=row_num, column=9, value=amc.get_status_display() if amc.status else 'Active')
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="amc_report.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def export_routine_service_csv(request):
    """Export Routine Services to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="routine_service_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Reference ID', 'Customer', 'AMC Type', 'Created', 
        'Start Date', 'End Date', 'Status'
    ])
    
    amc_records = AMC.objects.all().select_related('customer', 'amc_type')
    for amc in amc_records:
        writer.writerow([
            amc.reference_id or amc.id,
            amc.customer.site_name if amc.customer else 'N/A',
            amc.amc_type.name if amc.amc_type else 'N/A',
            amc.created.strftime('%Y-%m-%d %H:%M') if amc.created else '',
            amc.start_date.strftime('%Y-%m-%d') if amc.start_date else '',
            amc.end_date.strftime('%Y-%m-%d') if amc.end_date else '',
            amc.get_status_display() if amc.status else 'Active',
        ])
    
    return response


@login_required
def export_routine_service_xlsx(request):
    """Export Routine Services to XLSX"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Routine Service Report"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Reference ID', 'Customer', 'AMC Type', 'Created', 
        'Start Date', 'End Date', 'Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data
    amc_records = AMC.objects.all().select_related('customer', 'amc_type')
    for row_num, amc in enumerate(amc_records, 2):
        worksheet.cell(row=row_num, column=1, value=amc.reference_id or str(amc.id))
        worksheet.cell(row=row_num, column=2, value=amc.customer.site_name if amc.customer else 'N/A')
        worksheet.cell(row=row_num, column=3, value=amc.amc_type.name if amc.amc_type else 'N/A')
        created_cell = worksheet.cell(row=row_num, column=4)
        if amc.created:
            created_cell.value = amc.created
            created_cell.number_format = 'DD-MM-YYYY HH:MM'
        else:
            created_cell.value = ''
        start_cell = worksheet.cell(row=row_num, column=5)
        if amc.start_date:
            start_cell.value = amc.start_date
            start_cell.number_format = 'DD-MM-YYYY'
        else:
            start_cell.value = ''
        end_cell = worksheet.cell(row=row_num, column=6)
        if amc.end_date:
            end_cell.value = amc.end_date
            end_cell.number_format = 'DD-MM-YYYY'
        else:
            end_cell.value = ''
        worksheet.cell(row=row_num, column=7, value=amc.get_status_display() if amc.status else 'Active')
    
    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="routine_service_report.xlsx"'
    workbook.save(response)
    
    return response